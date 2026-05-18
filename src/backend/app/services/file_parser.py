from __future__ import annotations

import re
from pathlib import Path
from typing import List

from fastapi import UploadFile

from ..models.schemas import Chapter, Textbook
from ..utils.text import clean_line, clean_text, is_noise_line, looks_like_toc_page, safe_id

CHAPTER_RE = re.compile(
    r"^\s*((第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章)[\s\u3000：:、.-]*(.{0,40})|Chapter\s+\d+\s*[:.\-\s].{0,60})\s*$",
    re.IGNORECASE,
)
CHAPTER_PREFIX_RE = r"第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章"
CHAPTER_PREFIX_PATTERN = re.compile(CHAPTER_PREFIX_RE)
CHAPTER_TITLE_LINE_RE = re.compile(rf"^\s*(?:\d+\s+)?({CHAPTER_PREFIX_RE})(.*)$")
SECTION_RE = re.compile(r"^\s*(第\s*[一二三四五六七八九十百千万零〇两\d]+\s*节|[一二三四五六七八九十]+、|\d+[\.、])")
TITLE_STOP_MARKERS = ("本章数字资源", "学习目标", "本章思维导图", "推荐阅读", "中英文名词对照索引", "附录", "数字特色", "AR 模型")


class TextbookParser:
    def __init__(self, upload_dir: Path) -> None:
        self.upload_dir = upload_dir
        self.upload_dir.mkdir(parents=True, exist_ok=True)

    async def save_upload(self, file: UploadFile) -> Path:
        suffix = Path(file.filename or "upload.bin").suffix.lower()
        stem = Path(file.filename or "upload").stem
        target = self.upload_dir / f"{safe_id('file', stem + suffix)}{suffix}"
        with target.open("wb") as f:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                f.write(chunk)
        return target

    def parse(self, path: Path, original_name: str | None = None) -> Textbook:
        fmt = path.suffix.lower().lstrip(".")
        textbook_id = safe_id("book", f"{path.name}:{path.stat().st_size}")
        filename = original_name or path.name
        title = Path(filename).stem
        if fmt == "pdf":
            chapters, pages = self._parse_pdf(path)
        elif fmt in {"md", "markdown"}:
            chapters, pages = self._parse_markdown(path), 1
        elif fmt == "txt":
            chapters, pages = self._parse_txt(path), 1
        elif fmt == "docx":
            chapters, pages = self._parse_docx(path), 1
        elif fmt in {"xlsx", "xlsm"}:
            chapters, pages = self._parse_xlsx(path), 1
        else:
            raise ValueError(f"暂不支持的文件格式：{fmt}")

        total_chars = sum(c.char_count for c in chapters)
        return Textbook(
            textbook_id=textbook_id,
            filename=filename,
            title=title,
            format=fmt,
            file_size=path.stat().st_size,
            total_pages=pages,
            total_chars=total_chars,
            status="已完成",
            chapters=chapters,
        )

    def _parse_pdf(self, path: Path) -> tuple[List[Chapter], int]:
        import fitz  # PyMuPDF

        doc = fitz.open(path)
        page_count = doc.page_count
        chapters: List[Chapter] = []
        current_title = "未识别章节"
        current_start = 1
        current_pages: List[str] = []

        for page_index in range(doc.page_count):
            page = doc.load_page(page_index)
            page_no = page_index + 1
            text = self._extract_pdf_page_text(page)
            if not text or looks_like_toc_page(text):
                continue
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            title = self._detect_chapter_title(lines)
            if title and current_pages and self._chapter_key(title) != self._chapter_key(current_title):
                chapters.append(
                    self._make_chapter(len(chapters) + 1, current_title, current_start, page_no - 1, "\n".join(current_pages))
                )
                current_pages = []
                current_title = title
                current_start = page_no
            elif title and not current_pages:
                current_title = title
                current_start = page_no
            current_pages.append(text)

        if current_pages:
            chapters.append(self._make_chapter(len(chapters) + 1, current_title, current_start, doc.page_count, "\n".join(current_pages)))
        doc.close()

        chapters = self._postprocess_chapters(chapters)
        if len(chapters) <= 1:
            all_text = chapters[0].content if chapters else ""
            chapters = self._fallback_split(all_text)
        return chapters, page_count

    def _extract_pdf_page_text(self, page) -> str:
        # Read per page and filter header/footer + tiny figure/table fragments.
        blocks = page.get_text("blocks", sort=True)
        h = page.rect.height
        kept: List[str] = []
        for b in blocks:
            x0, y0, x1, y1, txt = b[:5]
            if y0 < h * 0.04 or y1 > h * 0.965:
                continue
            if not txt or len(txt.strip()) < 2:
                continue
            # Skip blocks that look like figure captions or table remnants, but keep chapter titles.
            striped = txt.strip()
            if len(striped) < 60 and re.search(r"^(图|表)\s*\d", striped):
                continue
            cleaned_lines = [clean_line(line) for line in striped.splitlines()]
            cleaned_lines = [line for line in cleaned_lines if not is_noise_line(line)]
            if cleaned_lines:
                kept.append("\n".join(cleaned_lines))
        return clean_text("\n".join(kept))

    def _detect_chapter_title(self, lines: List[str]) -> str | None:
        scan_lines = lines[:14]
        inferred = self._infer_chapter_title_from_lines(scan_lines)
        if inferred:
            return inferred
        for index, line in enumerate(scan_lines):
            compact = re.sub(r"\s+", " ", line).strip()
            if is_noise_line(compact) or "目录" in compact:
                continue
            candidates = [compact]
            if CHAPTER_PREFIX_PATTERN.fullmatch(compact) and index + 1 < len(scan_lines):
                next_line = re.sub(r"\s+", " ", scan_lines[index + 1]).strip()
                if next_line and not is_noise_line(next_line):
                    candidates.insert(0, f"{compact} {next_line}")
            for candidate in candidates:
                if CHAPTER_RE.match(candidate) and 3 <= len(candidate) <= 80:
                    return self._normalize_chapter_title(candidate)
        return None

    def _parse_markdown(self, path: Path) -> List[Chapter]:
        text = clean_text(path.read_text(encoding="utf-8", errors="ignore"))
        lines = text.splitlines()
        chapters: List[Chapter] = []
        title = "未命名章节"
        buf: List[str] = []
        for line in lines:
            if line.startswith("#") or CHAPTER_RE.match(line.strip()):
                if buf:
                    chapters.append(self._make_chapter(len(chapters) + 1, title, len(chapters) + 1, len(chapters) + 1, "\n".join(buf)))
                    buf = []
                title = re.sub(r"^#+\s*", "", line).strip()
            else:
                buf.append(line)
        if buf:
            chapters.append(self._make_chapter(len(chapters) + 1, title, len(chapters) + 1, len(chapters) + 1, "\n".join(buf)))
        return chapters or self._fallback_split(text)

    def _parse_txt(self, path: Path) -> List[Chapter]:
        text = clean_text(path.read_text(encoding="utf-8", errors="ignore"))
        lines = text.splitlines()
        chapters: List[Chapter] = []
        title = "未命名章节"
        buf: List[str] = []
        for line in lines:
            if CHAPTER_RE.match(line.strip()):
                if buf:
                    chapters.append(self._make_chapter(len(chapters) + 1, title, len(chapters) + 1, len(chapters) + 1, "\n".join(buf)))
                    buf = []
                title = line.strip()
            else:
                buf.append(line)
        if buf:
            chapters.append(self._make_chapter(len(chapters) + 1, title, len(chapters) + 1, len(chapters) + 1, "\n".join(buf)))
        return chapters or self._fallback_split(text)

    def _parse_docx(self, path: Path) -> List[Chapter]:
        from docx import Document

        doc = Document(str(path))
        lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        tmp = self.upload_dir / "_tmp_docx.txt"
        tmp.write_text("\n".join(lines), encoding="utf-8")
        return self._parse_txt(tmp)

    def _parse_xlsx(self, path: Path) -> List[Chapter]:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        pieces: List[str] = []
        for ws in wb.worksheets:
            pieces.append(f"# {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [str(v) for v in row if v is not None]
                if values:
                    pieces.append("\t".join(values))
        text = clean_text("\n".join(pieces))
        return self._fallback_split(text, default_title="Excel 数据表")

    def _make_chapter(self, idx: int, title: str, page_start: int, page_end: int, content: str) -> Chapter:
        content = clean_text(content)
        content, inferred_title = self._trim_to_chapter_body(content)
        clean_title = clean_line(title)
        if inferred_title and (
            "未识别" in clean_title
            or "未命名" in clean_title
            or clean_title.startswith("自动分段")
            or self._title_has_more_detail(inferred_title, clean_title)
            or self._title_is_suspicious(clean_title)
        ):
            clean_title = inferred_title
        return Chapter(
            chapter_id=f"ch_{idx:03d}",
            title=clean_title or f"第 {idx} 部分",
            page_start=max(1, page_start),
            page_end=max(page_start, page_end),
            content=content,
            char_count=len(content),
        )

    def _postprocess_chapters(self, chapters: List[Chapter]) -> List[Chapter]:
        merged: List[Chapter] = []
        for chapter in chapters:
            if not chapter.content or chapter.char_count < 40:
                continue
            if looks_like_toc_page(chapter.content):
                continue
            if self._looks_like_catalog_fragment(chapter.content):
                continue
            if self._looks_like_front_matter(chapter):
                continue
            key = self._chapter_key(chapter.title)
            if merged and key and key == self._chapter_key(merged[-1].title):
                previous = merged[-1]
                merged[-1] = self._make_chapter(
                    len(merged),
                    previous.title,
                    previous.page_start,
                    chapter.page_end,
                    f"{previous.content}\n{chapter.content}",
                )
            else:
                merged.append(chapter)
        return [
            self._make_chapter(idx, chapter.title, chapter.page_start, chapter.page_end, chapter.content)
            for idx, chapter in enumerate(merged, 1)
        ]

    @staticmethod
    def _chapter_key(title: str) -> str:
        title = clean_line(title)
        title = re.sub(r"\s+", "", title)
        match = re.match(r"^(第[一二三四五六七八九十百千万零〇两\d]+章)", title)
        return match.group(1) if match else title

    @staticmethod
    def _title_has_more_detail(candidate: str, current: str) -> bool:
        candidate = clean_line(candidate)
        current = clean_line(current)
        if not candidate or candidate == current:
            return False
        candidate_key = TextbookParser._chapter_key(candidate)
        current_key = TextbookParser._chapter_key(current)
        if candidate_key != current_key:
            return False
        bare = r"^第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章\s*$"
        truncated = r"^第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章\s*[上下中]$"
        return bool(re.match(bare, current) or re.match(truncated, current) or len(candidate) >= len(current) + 2)

    @staticmethod
    def _title_is_suspicious(title: str) -> bool:
        title = clean_line(title)
        if not title:
            return True
        if any(marker in title for marker in ["目录", "推荐阅读", "索引", "附录", "AR 模型", "数字特色"]):
            return True
        if re.search(r"[。！？；;]", title):
            return True
        match = CHAPTER_TITLE_LINE_RE.match(title)
        if match:
            rest = re.sub(r"\s+", "", match.group(2))
            if len(rest) > 22:
                return True
            if re.search(r"\d", rest):
                return True
        return False

    @staticmethod
    def _looks_like_catalog_fragment(content: str) -> bool:
        head = clean_text(content[:1800])
        if "本章数字资源" in head:
            return False
        chapter_hits = len(re.findall(r"第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章", head))
        back_matter_hits = sum(marker in head for marker in ["推荐阅读", "中英文名词对照索引", "附录", "数字特色", "AR 模型"])
        catalog_rows = len(re.findall(r"(?:第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章|推荐阅读|索引|附录|AR\s*模型).{0,24}\d+", head))
        if len(head) < 500 and chapter_hits >= 1:
            return True
        return chapter_hits >= 2 or back_matter_hits >= 2 or catalog_rows >= 2

    @staticmethod
    def _looks_like_front_matter(chapter: Chapter) -> bool:
        title = clean_line(chapter.title)
        head = clean_text(chapter.content[:1800])
        if "本章数字资源" in head:
            return False
        if CHAPTER_PREFIX_PATTERN.search(title) and "未识别" not in title and "未命名" not in title:
            return False
        publishing_hits = sum(marker in head for marker in ["规划教材", "人民卫生出版社", "主编", "副主编", "版权", "目录", "编写", "修订"])
        return "未识别" in title or "未命名" in title or title.startswith("自动分段") or publishing_hits >= 2

    @staticmethod
    def _trim_to_chapter_body(content: str) -> tuple[str, str | None]:
        inferred_from_text = TextbookParser._infer_chapter_title_from_content(content)
        lines = [line for line in content.splitlines() if line.strip()]
        resource_index = next((i for i, line in enumerate(lines) if "本章数字资源" in line), -1)
        if resource_index < 0:
            return content, inferred_from_text
        title_start = max(0, resource_index - 3)
        title_lines = [
            line.strip()
            for line in lines[title_start:resource_index]
            if not re.fullmatch(r"\d+", line.strip()) and not any(marker in line for marker in ["目录", "推荐阅读", "索引", "附录"])
        ]
        inferred_title = inferred_from_text or clean_line(re.sub(r"\s+", " ", "".join(title_lines)))
        if inferred_title:
            inferred_title = TextbookParser._normalize_chapter_title(inferred_title)
        should_trim = resource_index > 4 or any(marker in "\n".join(lines[:resource_index]) for marker in ["目录", "推荐阅读", "索引", "附录", "AR 模型"])
        if should_trim:
            content = "\n".join(lines[title_start:])
        if not (2 <= len(inferred_title) <= 40):
            inferred_title = None
        return content, inferred_title

    @staticmethod
    def _infer_chapter_title_from_content(content: str) -> str | None:
        head = clean_text(content[:2600])
        inferred = TextbookParser._infer_chapter_title_from_lines(head.splitlines()[:28])
        if inferred:
            return inferred
        match = re.search(r"(^|\s)(绪\s*论)\s*本章数字资源", head)
        if match:
            return "绪论"
        return None

    @staticmethod
    def _infer_chapter_title_from_lines(lines: List[str]) -> str | None:
        cleaned = []
        for line in lines[:30]:
            compact = clean_line(re.sub(r"\s+", " ", line)).strip()
            if not compact or "目录" in compact:
                continue
            if is_noise_line(compact) and not CHAPTER_PREFIX_PATTERN.search(compact):
                continue
            cleaned.append(compact)

        resource_index = next((i for i, line in enumerate(cleaned) if "本章数字资源" in line), -1)
        if resource_index >= 0:
            title = TextbookParser._extract_title_near_resource(cleaned, resource_index)
            if title:
                return title

        for index, line in enumerate(cleaned):
            match = CHAPTER_TITLE_LINE_RE.match(line)
            if not match:
                continue
            if TextbookParser._looks_like_title_catalog_row(line):
                continue

            prefix = match.group(1)
            pieces: List[str] = []
            tail = match.group(2).strip(" \u3000：:、.-|")
            tail = TextbookParser._clean_title_piece(tail)
            if tail:
                pieces.append(tail)

            for follow in cleaned[index + 1 : index + 6]:
                piece = TextbookParser._clean_title_piece(follow)
                if not piece:
                    continue
                if TextbookParser._is_title_stop_line(piece):
                    if pieces:
                        break
                    continue
                if CHAPTER_PREFIX_PATTERN.search(piece) or TextbookParser._looks_like_title_catalog_row(piece):
                    break
                if TextbookParser._looks_like_body_line(piece):
                    break
                pieces.append(piece)
                if len("".join(pieces)) >= 24:
                    break

            raw_title = f"{prefix} {''.join(pieces)}" if pieces else prefix
            title = TextbookParser._normalize_chapter_title(raw_title)
            if TextbookParser._is_valid_chapter_title(title):
                return title
        return None

    @staticmethod
    def _extract_title_near_resource(lines: List[str], resource_index: int) -> str | None:
        start = max(0, resource_index - 6)
        window = lines[start:resource_index]
        if not window:
            return None
        chapter_pos = -1
        chapter_match: re.Match[str] | None = None
        for idx, line in enumerate(window):
            match = CHAPTER_TITLE_LINE_RE.match(line)
            if match and not TextbookParser._looks_like_title_catalog_row(line):
                chapter_pos = idx
                chapter_match = match
        if chapter_match:
            pieces: List[str] = []
            tail = TextbookParser._clean_title_piece(chapter_match.group(2).strip(" \u3000：:、.-|"))
            if tail:
                pieces.append(tail)
            for follow in window[chapter_pos + 1 :]:
                piece = TextbookParser._clean_title_piece(follow)
                if not piece or TextbookParser._is_title_stop_line(piece):
                    continue
                if CHAPTER_TITLE_LINE_RE.match(piece) or TextbookParser._looks_like_body_line(piece):
                    break
                pieces.append(piece)
            title = TextbookParser._normalize_chapter_title(f"{chapter_match.group(1)} {''.join(pieces)}")
            return title if TextbookParser._is_valid_chapter_title(title) else None

        plain = "".join(
            TextbookParser._clean_title_piece(line)
            for line in window[-3:]
            if not re.fullmatch(r"\d+", line) and not TextbookParser._is_title_stop_line(line)
        )
        plain = re.sub(r"\s+", "", plain)
        if plain == "绪论":
            return "绪论"
        return None

    @staticmethod
    def _clean_title_piece(text: str) -> str:
        text = clean_line(text)
        text = re.sub(r"^[\s\u3000：:、.-]+|[\s\u3000：:、.-]+$", "", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    @staticmethod
    def _is_title_stop_line(line: str) -> bool:
        if re.fullmatch(r"\d+", line):
            return True
        if any(marker in line for marker in TITLE_STOP_MARKERS):
            return True
        return bool(SECTION_RE.match(line))

    @staticmethod
    def _looks_like_body_line(line: str) -> bool:
        if len(line) > 30:
            return True
        if re.search(r"[。！？；;]", line):
            return True
        return False

    @staticmethod
    def _looks_like_title_catalog_row(line: str) -> bool:
        return bool(re.search(r"(?:第\s*[一二三四五六七八九十百千万零〇两\d]+\s*章|推荐阅读|索引|附录|AR\s*模型).{0,24}\d+\s*$", line))

    @staticmethod
    def _is_valid_chapter_title(title: str) -> bool:
        if not title or len(title) > 45:
            return False
        if any(marker in title for marker in ["目录", "推荐阅读", "索引", "附录", "AR 模型", "数字特色"]):
            return False
        if re.search(r"[。！？；;]", title):
            return False
        return bool(CHAPTER_PREFIX_PATTERN.match(title))

    @staticmethod
    def _normalize_chapter_title(title: str) -> str:
        title = clean_line(title)
        title = re.sub(r"\s+", " ", title).strip()
        title = re.sub(r"(第)\s*([一二三四五六七八九十百千万零〇两\d]+)\s*(章)", r"\1\2\3 ", title)
        title = re.split(r"\s*(?:本章数字资源|学习目标|本章思维导图|第一节|第二节|一、|二、)", title, maxsplit=1)[0]
        title = re.sub(r"(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])", "", title)
        title = re.sub(r"(第[一二三四五六七八九十百千万零〇两\d]+章)(.+?)\s+\2.+$", r"\1\2", title)
        title = re.sub(r"\s+", " ", title).strip(" ：:、.-")
        title = re.sub(r"(第[一二三四五六七八九十百千万零〇两\d]+章)\s*", r"\1 ", title).strip()
        return title

    def _fallback_split(self, text: str, default_title: str = "自动分段") -> List[Chapter]:
        size = 16000
        chapters: List[Chapter] = []
        for idx, start in enumerate(range(0, len(text), size), 1):
            piece = text[start : start + size]
            chapters.append(self._make_chapter(idx, f"{default_title} {idx}", idx, idx, piece))
        return chapters
